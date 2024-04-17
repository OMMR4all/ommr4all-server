
from database.file_formats.pcgts import SymbolType

if __name__ == '__main__':
    import openpyxl

    from database import DatabaseBook
    from database.database_book_documents import DatabaseBookDocuments

def integrate_and_algin_sheet_in_workbook(sheet, sheet_start_row, stat_sheet, documents):
    for i in range(sheet_start_row, 2500):
        doc = documents.database_documents.get_document_by_id(sheet["A" + str(i)].value)
        print(doc)
        if doc:
            symbols, meta = doc.get_symbols(book=c)
            n_symbols = 0
            n_note_components = 0
            n_clefs = 0
            n_accids = 0
            for t in symbols:
                n_symbols += len(t)
                n_note_components += len([s for s in t if s.symbol_type == SymbolType.NOTE])
                n_clefs += len([s for s in t if s.symbol_type == SymbolType.CLEF])
                n_accids += len([s for s in t if s.symbol_type == SymbolType.ACCID])

            stat_sheet.cell(row=i, column=1, value=sheet["A" + str(i)].value)
            stat_sheet.cell(row=i, column=2, value=n_symbols)
            stat_sheet.cell(row=i, column=3, value=n_note_components)
            stat_sheet.cell(row=i, column=4, value=n_clefs)
            stat_sheet.cell(row=i, column=5, value=n_accids)




    return None

c = DatabaseBook('Graduel_Syn22_03_24')
documents = DatabaseBookDocuments().load(c)
workbook = openpyxl.load_workbook('/tmp/Graduale_w_stats2.xlsx')
sheet = workbook.get_sheet_by_name("Sheet1")
for i in range(3, 2500):
    doc = documents.database_documents.get_document_by_id(sheet["A" + str(i)].value)
    lines = 0
    print(doc)
    if doc:
        symbols, meta = doc.get_symbols(book=c)
        n_symbols = 0
        n_note_components = 0
        n_clefs = 0
        n_accids = 0
        for t in symbols:
            n_symbols += len(t)
            n_note_components += len([s for s in t if s.symbol_type == SymbolType.NOTE])
            n_clefs += len([s for s in t if s.symbol_type == SymbolType.CLEF])
            n_accids += len([s for s in t if s.symbol_type == SymbolType.ACCID])
        lines = len(doc.pages_ids)
    sheet["AT" + str(i)].value = lines
workbook.save('/tmp/Graduale_comb_stats_pages.xlsx')

