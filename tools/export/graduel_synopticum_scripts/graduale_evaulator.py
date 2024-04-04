import os
from typing import List

import openpyxl

from database import DatabaseBook
from database.database_book_documents import DatabaseBookDocuments, DocSpanType
from omr.experimenter.documents.b_evalluator import evaluate_symbols, evaluate_syllabels
from omr.experimenter.documents.doc_instace_evaluator import write_staffline_eval_data, eval_layout_docs_instance, \
    eval_layout_docs_line_instance, eval_symbols__docs_instance, eval_texts_docs_instance, eval_texts_line_instance, \
    eval_texts_docs_instance_ignore_lines, eval_syl_docs_instance, eval_syl_docs_line_instance, \
    gen_eval_symbol_documents_data, gen_eval_syllable_documents_data
from omr.experimenter.documents.evaluater import evaluate_stafflines


def overview_excel_sheet(symbol_eval_data, syl_eval_data, sheet):
    pred, gt = symbol_eval_data.get_symbol_data()
    excel_symbol = evaluate_symbols(pred, gt)

    ex_syll, errors = evaluate_syllabels(syl_eval_data.get_syl_data())
    ind = 3
    left = 3
    for t in [excel_symbol, ex_syll]:

        for ind_d, d in enumerate(t):

            for ind1, cell in enumerate(d):
                sheet.write(ind, ind1 + left, cell)
            ind += 1
    pass


from ommr4all import settings


def prepare_documents(gt_book: DatabaseBook):
    workbook = openpyxl.load_workbook('/tmp/Graduale1.xlsx')
    sheet = workbook.active
    documents = DatabaseBookDocuments().load(gt_book)
    docs: List[DocSpanType] = []

    for i in range(2, 2500):
        doc = documents.database_documents.get_document_by_monodi_id(sheet["AD" + str(i)].value)
        print(doc)
        print(i)
        if doc:
            docs.append(DocSpanType(p_start=doc.start.page_name, p_end=doc.end.page_name, doc=doc, index=documents.database_documents.documents.index(doc)))

    return docs


books = ['Graduel_Syn']
for i in books:
    b = DatabaseBook(i)
    c = DatabaseBook('Graduel_Syn22_03_24')

    # excel_lines1 = evaluate_stafflines(b, c)
    docs = prepare_documents(c)
    #docs = filter_docs(docs, 253)
    symbol_eval_data = gen_eval_symbol_documents_data(b, c, docs)
    syl_eval_data = gen_eval_syllable_documents_data(b, c, docs)

    from xlwt import Workbook

    wb2 = Workbook()
    # Workbook is created
    wb = Workbook()
    # add_sheet is used to create sheet.

    sheet1 = wb.add_sheet('Symbols Docs')
    eval_symbols__docs_instance(symbol_eval_data, sheet1)
    sheet2 = wb.add_sheet('Symbols Lines')
    eval_symbols__docs_instance(symbol_eval_data, sheet2)

    sheet5 = wb.add_sheet('Syllable Docs')
    eval_syl_docs_instance(syl_eval_data, sheet5)
    sheet6 = wb.add_sheet('Syllable Line')
    eval_syl_docs_line_instance(syl_eval_data, sheet6)
    sheet10 = wb.add_sheet('overview 1')
    overview_excel_sheet(symbol_eval_data, syl_eval_data, sheet10)

    wb.save(f"/tmp/{i}.xls")
