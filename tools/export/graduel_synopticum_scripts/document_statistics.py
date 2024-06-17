from database.file_formats.pcgts import SymbolType

if __name__ == '__main__':
    import openpyxl

    from database import DatabaseBook
    from database.database_book_documents import DatabaseBookDocuments

def integrate_and_algin_sheet_in_workbook(sheet, sheet_start_row, stat_sheet, documents):
    for i in range(sheet_start_row, 2500):
        doc = documents.database_documents.get_document_by_id(sheet["A" + str(i)].value)
        print(doc)
        if doc:
            symbols, meta = doc.get_symbols(book=c)
            n_symbols = 0
            n_note_components = 0
            n_clefs = 0
            n_accids = 0
            syllabels = 0
            chars = 0
            for t in symbols:
                n_symbols += len(t)
                n_note_components += len([s for s in t if s.symbol_type == SymbolType.NOTE])
                n_clefs += len([s for s in t if s.symbol_type == SymbolType.CLEF])
                n_accids += len([s for s in t if s.symbol_type == SymbolType.ACCID])
            lines = doc.get_page_line_of_document(c)
            for l, page in lines:
                syllabels += len(l.sentence.syllables)
                chars += len(l.text())
            stat_sheet.cell(row=i, column=1, value=sheet["A" + str(i)].value)
            stat_sheet.cell(row=i, column=2, value=n_symbols)
            stat_sheet.cell(row=i, column=3, value=n_note_components)
            stat_sheet.cell(row=i, column=4, value=n_clefs)
            stat_sheet.cell(row=i, column=5, value=n_accids)
            stat_sheet.cell(row=i, column=6, value=syllabels)
            stat_sheet.cell(row=i, column=7, value=chars)




    return None


workbook = openpyxl.load_workbook('/tmp/Graduale.xlsx')
sheet = workbook.get_sheet_by_name("Sheet1")



sheet2= workbook.create_sheet("Symbols Doc Statistics")

c = DatabaseBook('Graduel_Syn22_03_24')
documents = DatabaseBookDocuments().load(c)

integrate_and_algin_sheet_in_workbook(sheet, 2, sheet2, documents)

"""
documents = DatabaseBookDocuments().load(c)
docs = {}
for i in range(5, 2500):
    doc = documents.database_documents.get_document_by_id(sheet_a["A" + str(i)].value)
    if doc:
        print(i)

        docs[doc.monody_id] = i
rows = list(sheet_a.rows)
for i in range(2, 2500):
    x = docs.get(sheet["AD" + str(i)].value, None)
    sheet3["A" + str(i)] = sheet["AD" + str(i)].value
    if x:
        row = rows[x]
        for ind, cell in enumerate(row):
            sheet3.cell(row=i, column=ind + 2, value=cell.value)
    #print(x)
    #doc = documents.database_documents.get_document_by_monodi_id(sheet["AD" + str(i)].value)
"""
workbook.save("/tmp/Graduale_stats.xlsx")