import openpyxl

from database import DatabaseBook
from database.database_book_documents import DatabaseBookDocuments

if __name__ == "__main__":

    def integrate_and_algin_sheet_in_workbook(sheet, sheet_start_row,  sheet_a, sheet_a_start_row, new_sheet, documents):
        docs = {}
        for i in range(sheet_a_start_row, 2500):
            doc = documents.database_documents.get_document_by_id(sheet_a["A" + str(i)].value)
            if doc:
                print(i)

                docs[doc.doc_id] = i
        rows = list(sheet_a.rows)
        for i in range(sheet_start_row, 2500):
            x = docs.get(sheet["A" + str(i)].value, None)
            if x:
                new_sheet["A" + str(i)] = sheet["A" + str(i)].value
                #new_sheet["A" + str(i)] =
                row = rows[x-1]
                for ind, cell in enumerate(row):
                    new_sheet.cell(row=i, column=ind + 2, value=cell.value)
        return sheet3
    #workbook = openpyxl.load_workbook('/tmp/Graduale_stats.xlsx')
    workbook = openpyxl.load_workbook('/tmp/Graduale.xlsx')

    sheet = workbook.active
    workbook2 = openpyxl.load_workbook('/tmp/Graduel_Syn22_03_24_pred.xlsx')

    sheet_a = workbook2.get_sheet_by_name("Symbols Docs")
    sheet_b = workbook2.get_sheet_by_name("Syllable Docs")

    sheet3 = workbook.create_sheet("Results Symbols1")
    sheet4 = workbook.create_sheet("Results Syllables1")

    c = DatabaseBook('Graduel_Syn22_03_24')
    documents = DatabaseBookDocuments().load(c)

    integrate_and_algin_sheet_in_workbook(sheet, 3, sheet_a, 6, sheet3, documents)
    integrate_and_algin_sheet_in_workbook(sheet, 3, sheet_b, 5, sheet4, documents)

    """
    documents = DatabaseBookDocuments().load(c)
    docs = {}
    for i in range(5, 2500):
        doc = documents.database_documents.get_document_by_id(sheet_a["A" + str(i)].value)
        if doc:
            print(i)

            docs[doc.monody_id] = i
    rows = list(sheet_a.rows)
    for i in range(2, 2500):
        x = docs.get(sheet["AD" + str(i)].value, None)
        sheet3["A" + str(i)] = sheet["AD" + str(i)].value
        if x:
            row = rows[x]
            for ind, cell in enumerate(row):
                sheet3.cell(row=i, column=ind + 2, value=cell.value)
        #print(x)
        #doc = documents.database_documents.get_document_by_monodi_id(sheet["AD" + str(i)].value)
    """
    workbook.save("/tmp/Graduale_comb.xlsx")
