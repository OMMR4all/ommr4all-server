import json
import os

import openpyxl

from database import DatabaseBook, DatabasePage, DatabaseFile
from database.database_book_documents import DatabaseBookDocuments
from database.file_formats.exporter.monodi.monodi2_exporter import PcgtsToMonodiConverter

def soruce_meta(id):
    meta = {
        "id": id,
        "quellensigle": "",
        "herkunftsregion": "",
        "herkunftsort": "",
        "herkunftsinstitution": "",
        "ordenstradition": "",
        "quellentyp": "",
        "bibliotheksort": "",
        "bibliothek": "",
        "bibliothekssignatur": "",
        "kommentar": "",
        "datierung": "",
        "status": "",
        "jahrhundert": "",
        "manifest": "",
        "foliooffset": 0,
        "publish": ""
    }
    return meta

if __name__ == "__main__":
    source = "OMMR_Test1"
    source_meta = soruce_meta(source)
    book = DatabaseBook('mul_2_rsync_gt')
    documents = DatabaseBookDocuments().load(book)
    docs = documents.database_documents.documents
    wb_obj = openpyxl.load_workbook("/tmp/CM Default Metadatendatei.xlsx")
    sheet_obj = wb_obj.active
    #cell_obj = sheet_obj.cell(row=1, column=30)
    #print(cell_obj.value)
    i = 2
    os.mkdir("/tmp/export/")
    os.mkdir(f"/tmp/export/{source}")
    with open(f"/tmp/export/{source}/meta.json", "w") as f:
        json.dump(source_meta, f, indent=4)
    while True:
        cell_obj = sheet_obj.cell(row=i, column=30).value
        if cell_obj:
            try:
                document = documents.database_documents.get_document_by_monodi_id(cell_obj)
                pages = [DatabasePage(book, x) for x in document.pages_names]
                pcgts = [DatabaseFile(page, 'pcgts', create_if_not_existing=True).page.pcgts() for page in pages]
                root = PcgtsToMonodiConverter(pcgts, document=document)
                meta, nodes = root.get_meta_and_notes(document=document, editor=str("OMMR4all"))
                os.mkdir(f"/tmp/export/{source}/{document.monody_id}")
                with open(f"/tmp/export/{source}/{document.monody_id}/data.json", "w") as f:
                    json.dump(nodes, f, indent=4)
                with open(f"/tmp/export/{source}/{document.monody_id}/meta.json", "w") as f:
                    json.dump(meta, f, indent=4)
            except:
                print(1)
            pass

        else:
            break
        i = i+1

    #docs_ids = []

    #pages = [DatabasePage(book, x) for x in document.pages_names]
    #pcgts = [DatabaseFile(page, 'pcgts', create_if_not_existing=True).page.pcgts() for page in
    #         pages]
    #root = PcgtsToMonodiConverter(pcgts, document=document)
    #json_data = root.get_Monodi_json(document=document, editor=str(editor))
